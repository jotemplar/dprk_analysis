"""Export Russian OSINT search results to Excel with search term attribution"""

import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from database.connection import get_session
from database.russian_search_models import RussianSearch, RussianSearchResult


def fetch_all_results(session: Session) -> List[Dict]:
    """
    Fetch all search results with query metadata

    Args:
        session: Database session

    Returns:
        List of result dictionaries with full metadata
    """
    results = session.query(RussianSearchResult).options(
        joinedload(RussianSearchResult.search)
    ).order_by(
        RussianSearchResult.search_id,
        RussianSearchResult.position
    ).all()

    data = []
    for result in results:
        search = result.search
        data.append({
            'query_id': search.query_id,
            'query_text': search.query_text,
            'engine': search.engine,
            'location': search.location,
            'language': search.language,
            'theme': search.theme,
            'sector': search.sector,
            'region': search.region,
            'time_filter': search.time_filter,
            'site': search.site,
            'position': result.position,
            'url': result.url,
            'title': result.title,
            'snippet': result.snippet,
            'source_domain': result.source_domain,
            'published_date': result.published_date.strftime('%Y-%m-%d') if result.published_date else '',
            'searched_at': search.searched_at.strftime('%Y-%m-%d %H:%M') if search.searched_at else ''
        })

    return data


def create_excel_report(data: List[Dict], output_path: str):
    """
    Create Excel report with formatted results

    Args:
        data: List of result dictionaries
        output_path: Path to save Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Russian OSINT Results"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_style = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Define headers
    headers = [
        'Query ID',
        'Engine',
        'Location',
        'Language',
        'Theme',
        'Sector',
        'Region',
        'Time Filter',
        'Site',
        'Query Text',
        'Position',
        'URL',
        'Title',
        'Snippet',
        'Source Domain',
        'Published Date',
        'Searched At'
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style

    # Write data
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=item['query_id'])
        ws.cell(row=row_num, column=2, value=item['engine'])
        ws.cell(row=row_num, column=3, value=item['location'])
        ws.cell(row=row_num, column=4, value=item['language'])
        ws.cell(row=row_num, column=5, value=item['theme'])
        ws.cell(row=row_num, column=6, value=item['sector'])
        ws.cell(row=row_num, column=7, value=item['region'])
        ws.cell(row=row_num, column=8, value=item['time_filter'])
        ws.cell(row=row_num, column=9, value=item['site'])
        ws.cell(row=row_num, column=10, value=item['query_text'])
        ws.cell(row=row_num, column=11, value=item['position'])

        # URL as hyperlink
        url_cell = ws.cell(row=row_num, column=12, value=item['url'])
        url_cell.hyperlink = item['url']
        url_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=row_num, column=13, value=item['title'])
        ws.cell(row=row_num, column=14, value=item['snippet'])
        ws.cell(row=row_num, column=15, value=item['source_domain'])
        ws.cell(row=row_num, column=16, value=item['published_date'])
        ws.cell(row=row_num, column=17, value=item['searched_at'])

        # Apply borders
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border_style

    # Adjust column widths
    column_widths = {
        'A': 25,  # Query ID
        'B': 15,  # Engine
        'C': 12,  # Location
        'D': 10,  # Language
        'E': 20,  # Theme
        'F': 20,  # Sector
        'G': 25,  # Region
        'H': 15,  # Time Filter
        'I': 20,  # Site
        'J': 60,  # Query Text
        'K': 10,  # Position
        'L': 50,  # URL
        'M': 50,  # Title
        'N': 60,  # Snippet
        'O': 25,  # Source Domain
        'P': 15,  # Published Date
        'Q': 18   # Searched At
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Apply filters
    ws.auto_filter.ref = ws.dimensions

    # Save workbook
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


def create_summary_sheet(wb: Workbook, session: Session):
    """
    Create summary statistics sheet

    Args:
        wb: Workbook object
        session: Database session
    """
    ws = wb.create_sheet("Summary", 0)

    # Get statistics
    total_queries = session.query(RussianSearch).count()
    completed_queries = session.query(RussianSearch).filter_by(search_status='completed').count()
    total_results = session.query(RussianSearchResult).count()

    yandex_queries = session.query(RussianSearch).filter_by(engine='yandex').count()
    google_queries = session.query(RussianSearch).filter_by(engine='google').count()

    # Write summary
    ws['A1'] = "Russian OSINT Search Results - Summary"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "Generated:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    ws['A5'] = "Total Queries:"
    ws['B5'] = total_queries

    ws['A6'] = "Completed Queries:"
    ws['B6'] = completed_queries

    ws['A7'] = "Total Results:"
    ws['B7'] = total_results

    ws['A9'] = "Yandex Queries:"
    ws['B9'] = yandex_queries

    ws['A10'] = "Google Russia Queries:"
    ws['B10'] = google_queries

    # Style
    for row in range(3, 11):
        ws[f'A{row}'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def export_to_excel(output_filename: str = None):
    """
    Main export function

    Args:
        output_filename: Custom output filename (optional)
    """
    session = get_session()

    try:
        print("Fetching search results from database...")
        data = fetch_all_results(session)

        if not data:
            print("No results found in database")
            return

        print(f"Found {len(data)} results")

        # Generate output filename
        if not output_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"reports/russian_osint_results_{timestamp}.xlsx"

        # Create Excel report
        print("Creating Excel report...")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary sheet first
        create_summary_sheet(wb, session)

        # Create results sheet
        ws = wb.create_sheet("Search Results")

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Define headers
        headers = [
            'Query ID', 'Engine', 'Location', 'Language', 'Theme',
            'Sector', 'Region', 'Time Filter', 'Site', 'Query Text',
            'Position', 'URL', 'Title', 'Snippet', 'Source Domain',
            'Published Date', 'Searched At'
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # Write data
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=item['query_id'])
            ws.cell(row=row_num, column=2, value=item['engine'])
            ws.cell(row=row_num, column=3, value=item['location'])
            ws.cell(row=row_num, column=4, value=item['language'])
            ws.cell(row=row_num, column=5, value=item['theme'])
            ws.cell(row=row_num, column=6, value=item['sector'])
            ws.cell(row=row_num, column=7, value=item['region'])
            ws.cell(row=row_num, column=8, value=item['time_filter'])
            ws.cell(row=row_num, column=9, value=item['site'])
            ws.cell(row=row_num, column=10, value=item['query_text'])
            ws.cell(row=row_num, column=11, value=item['position'])

            # URL as hyperlink
            url_cell = ws.cell(row=row_num, column=12, value=item['url'])
            if item['url']:
                url_cell.hyperlink = item['url']
                url_cell.font = Font(color="0563C1", underline="single")

            ws.cell(row=row_num, column=13, value=item['title'])
            ws.cell(row=row_num, column=14, value=item['snippet'])
            ws.cell(row=row_num, column=15, value=item['source_domain'])
            ws.cell(row=row_num, column=16, value=item['published_date'])
            ws.cell(row=row_num, column=17, value=item['searched_at'])

            # Apply borders
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border_style

        # Adjust column widths
        column_widths = [25, 15, 12, 10, 20, 20, 25, 15, 20, 60, 10, 50, 50, 60, 25, 15, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Apply filters
        ws.auto_filter.ref = ws.dimensions

        # Save workbook
        wb.save(output_filename)
        print(f"\nExcel report saved to: {output_filename}")
        print(f"Total results exported: {len(data)}")

    finally:
        session.close()


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='Export Russian OSINT search results to Excel')
    parser.add_argument('--output', type=str, help='Output filename')

    args = parser.parse_args()

    export_to_excel(args.output)


if __name__ == '__main__':
    main()
